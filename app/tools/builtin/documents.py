"""
文档处理工具：read_pdf / read_excel / extract_web。

解决：Agent 需要处理真实文档（论文 PDF、Excel 数据表、网页正文）。

    - read_pdf:     提取 PDF 文本（pymupdf，页级 + 总页数）
    - read_excel:   读取 Excel/CSV 表格（openpyxl，支持多 sheet + 行数限制）
    - extract_web:  抓取网页正文转 markdown 风格文本（比 http_get 更干净，限 16KB）

安全：
    - PDF/Excel 只能读沙箱内的文件（路径校验同 file_read）
    - 输出大小限制（防刷屏）
    - 网页抓取超时 + 大小限制
"""
import csv
import json
import re
from pathlib import Path

from pydantic import BaseModel, Field

from app.config import Settings
from app.errors import ToolExecutionError

# 输出限制
MAX_PDF_CHARS = 8000
MAX_EXCEL_CHARS = 8000
MAX_WEB_CHARS = 16000

_HTML_BLOCK_TAGS = ["p", "div", "section", "article", "h1", "h2", "h3", "h4", "h5", "li", "tr", "br"]


def _resolve_sandbox_path(relative_path: str) -> Path:
    """解析沙箱内路径（防穿越）。"""
    settings = Settings()
    sandbox = Path(settings.sandbox_dir).resolve()
    target = (sandbox / relative_path).resolve()
    if not str(target).startswith(str(sandbox)):
        raise ToolExecutionError(f"路径越界：只允许沙箱目录内（{sandbox}）")
    if not target.exists():
        raise ToolExecutionError(f"文件不存在: {relative_path}")
    return target


# ---------------------------------------------------------------------------
# read_pdf
# ---------------------------------------------------------------------------
class ReadPdfArgs(BaseModel):
    """PDF 读取参数。"""

    path: str = Field(description="沙箱内的 PDF 文件相对路径")
    page: int | None = Field(default=None, description="指定页（1 起）；缺省读全部（限 8KB）")


def read_pdf_handler(path: str, page: int | None = None) -> str:
    """提取 PDF 文本。"""
    target = _resolve_sandbox_path(path)
    if target.suffix.lower() != ".pdf":
        raise ToolExecutionError(f"不是 PDF 文件: {path}")

    try:
        import fitz  # PyMuPDF
    except ImportError as exc:
        raise ToolExecutionError("PyMuPDF 未安装，无法读取 PDF") from exc

    try:
        doc = fitz.open(target)
    except Exception as exc:
        raise ToolExecutionError(f"PDF 打开失败: {exc}") from exc

    try:
        total = doc.page_count
        if page is not None:
            if page < 1 or page > total:
                raise ToolExecutionError(f"页号越界: {page}（共 {total} 页）")
            pages = [page - 1]
        else:
            pages = list(range(total))

        lines = [f"📄 {path}（共 {total} 页）"]
        chars = 0
        for pno in pages:
            text = doc[pno].get_text().strip()
            if not text:
                continue
            if page is None:
                lines.append(f"\n--- 第 {pno + 1} 页 ---")
            lines.append(text)
            chars += len(text)
            if chars > MAX_PDF_CHARS:
                lines.append(f"\n...（内容截断，已输出 {chars} 字符）")
                break
        return "\n".join(lines)[: MAX_PDF_CHARS + 200]
    finally:
        doc.close()


# ---------------------------------------------------------------------------
# read_excel
# ---------------------------------------------------------------------------
class ReadExcelArgs(BaseModel):
    """Excel 读取参数。"""

    path: str = Field(description="沙箱内的 Excel（.xlsx/.xlsm）或 CSV 文件相对路径")
    sheet: str | None = Field(default=None, description="工作表名（缺省第一个）；CSV 忽略")
    max_rows: int = Field(default=20, ge=1, le=200, description="最大读取行数（1-200）")


def read_excel_handler(path: str, sheet: str | None = None, max_rows: int = 20) -> str:
    """读取 Excel/CSV 为文本表格。"""
    target = _resolve_sandbox_path(path)
    suffix = target.suffix.lower()

    # CSV：直接解析
    if suffix == ".csv":
        try:
            with open(target, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = [r for r in reader if any(c.strip() for c in r)]
        except (csv.Error, OSError) as exc:
            raise ToolExecutionError(f"CSV 读取失败: {exc}") from exc
        return _render_table(rows[: max_rows + 1], f"📊 {path}（{len(rows)} 行）")

    # Excel：openpyxl
    if suffix not in (".xlsx", ".xlsm"):
        raise ToolExecutionError(f"不支持的文件类型: {suffix}（支持 .xlsx/.xlsm/.csv）")
    try:
        import openpyxl
    except ImportError as exc:
        raise ToolExecutionError("openpyxl 未安装，无法读取 Excel") from exc

    try:
        wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
    except Exception as exc:
        raise ToolExecutionError(f"Excel 打开失败: {exc}") from exc

    try:
        sheets = wb.sheetnames
        ws = wb[sheet] if sheet else wb[sheets[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                rows.append(["" if v is None else str(v) for v in row])
            if len(rows) > max_rows:
                break
        total = ws.max_row or len(rows)
        header = f"📊 {path}（sheet: {ws.title}，共 {total} 行；sheets: {', '.join(sheets)}）"
        return _render_table(rows, header)
    finally:
        wb.close()


def _render_table(rows: list[list[str]], header: str) -> str:
    """把行数据渲染为 markdown 表格。"""
    if not rows:
        return header + "\n（无数据）"
    lines = [header, ""]
    # 表头 = 第一行
    header_row = rows[0]
    lines.append("| " + " | ".join(str(c)[:20] for c in header_row) + " |")
    lines.append("|" + "---|" * len(header_row))
    for r in rows[1 : min(len(rows), 21)]:
        lines.append("| " + " | ".join(str(c)[:30] for c in r) + " |")
    text = "\n".join(lines)
    return text[:MAX_EXCEL_CHARS]


# ---------------------------------------------------------------------------
# extract_web（网页正文 → markdown 风格）
# ---------------------------------------------------------------------------
class ExtractWebArgs(BaseModel):
    """网页正文提取参数。"""

    url: str = Field(description="要提取正文的网页地址（http/https）")


def extract_web_handler(url: str) -> str:
    """抓取网页并提取正文（去导航/脚本/样式，保留段落结构）。"""
    import httpx

    url = url.strip()
    if not url.startswith(("http://", "https://")):
        raise ToolExecutionError("只支持 http/https 协议")
    try:
        with httpx.Client(
            timeout=Settings().http_tool_timeout_seconds,
            follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (compatible; agent-runtime/0.1)"},
        ) as client:
            resp = client.get(url)
    except httpx.HTTPError as exc:
        raise ToolExecutionError(f"HTTP 请求失败: {exc}", transient=True) from exc
    if resp.status_code != 200:
        raise ToolExecutionError(f"HTTP {resp.status_code}")

    html = resp.text[:200000]  # 预截断原始 HTML
    # 移除脚本/样式/注释
    html = re.sub(r"<(script|style|noscript|svg|nav|footer|header)[\s\S]*?</\1>", " ", html, flags=re.I)
    html = re.sub(r"<!--[\s\S]*?-->", " ", html)
    # 块级标签转换行
    html = re.sub(r"<(" + "|".join(_HTML_BLOCK_TAGS) + r")[^>]*>", "\n", html, flags=re.I)
    # 其余标签去掉
    text = re.sub(r"<[^>]+>", "", html)
    # 清理空白
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s*\n+", "\n\n", text).strip()

    if not text:
        return "(页面无可提取正文)"
    if len(text) > MAX_WEB_CHARS:
        text = text[:MAX_WEB_CHARS] + f"\n...（截断，共 {len(text)} 字符）"
    return f"📄 {url}\n\n{text}"
